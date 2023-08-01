
import datetime
from helpers import SqlConnect
from openpyxl import *
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from models.raporlar.mkRaporlari import * 
import shutil

class MkRaporlari:
    def __init__(self):
        self.data = SqlConnect().data
        self.poBazindaNavlun = []
        self.marketingNavlun = []
        self.navlunYukleme = []
        self.marketingYuklemeNavlun = []
        self.yuklenenMusteriMasraf = []
    def getPoBazindaYillikSiparisler(self,yil):
        try:
            result = self.data.getStoreList("""
                                                select 
                                                    s.SiparisNo,
                                                    m.FirmaAdi,
                                                    sum(su.SatisToplam) as SiparisToplam,
													st.TeslimTur,
                                                    s.SiparisTarihi
                                                from SiparislerTB s
                                                    inner join SiparisUrunTB su on su.SiparisNo = s.SiparisNo
                                                    inner join MusterilerTB m on m.ID = s.MusteriID
													inner join SiparisTeslimTurTB st on st.ID = s.TeslimTurID
                                                where YEAR(s.SiparisTarihi)= ? and m.Marketing = 'Mekmar'

                                                group by
                                                    s.SiparisNo,m.FirmaAdi,st.TeslimTur,s.SiparisTarihi
                                                order by sum(su.SatisToplam) desc
                                            """,(yil))
            self.poBazindaNavlun = self.data.getStoreList("""
                                                            select 
                                                                s.SiparisNo,
                                                                m.FirmaAdi,
                                                                s.NavlunSatis + s.DetayTutar_1 + s.DetayTutar_2 + s.DetayTutar_3 + s.DetayTutar_4 as GelenTotal

                                                            from SiparislerTB s
                                                                inner join MusterilerTB m on m.ID = s.MusteriID
                                                            where YEAR(s.SiparisTarihi)= ? and m.Marketing = 'Mekmar'
                                                          
                                                          """,(yil))
            liste = list()
            for item in result:
                model = PoBazindaYillikModel()
                model.po = item.SiparisNo
                model.fob = self.__getNoneType(item.SiparisToplam)
                model.ddp = model.fob + self.__getNoneType(self.__getPoBazindaNavlun(item.SiparisNo))
                model.teslim = item.TeslimTur
                model.firma = item.FirmaAdi
                model.tarih = item.SiparisTarihi
                liste.append(model)
            schema = PoBazindaYillikSchema(many=True)
            return schema.dump(liste)
            
        except Exception as e:
            print('getPoBazindaYillikSiparisler hata',str(e))
            return False
        
    def __getPoBazindaNavlun(self,po):
        for item in self.poBazindaNavlun:
            if(item.SiparisNo == po):
                return item.GelenTotal
    
    def getMusteriBazindaUretim(self,yil):
        result = self.data.getStoreList("""
                                            select            
                                                    m.ID as MusteriId,            
                                                    m.FirmaAdi as MusteriAdi,               
                                                    m.Marketing,
                                                    (select yu.UlkeAdi from YeniTeklif_UlkeTB yu where yu.Id = m.UlkeId ) as Ulke,  
                                            (          
                                                Select Sum(SatisToplam) from SiparislerTB s, SiparisUrunTB u where s.SiparisNo=u.SiparisNo and s.SiparisDurumID=2 and s.MusteriID=m.ID and Year(s.SiparisTarihi)=?          
                                                
                                            ) as FOB,
                                            (
                                                select sum(s.NavlunSatis) + sum(DetayTutar_1) + sum(DetayTutar_2) + sum(DetayTutar_3) + sum(DetayTutar_4) from SiparislerTB s where s.SiparisDurumID=2 and s.MusteriID = m.ID and YEAR(s.SiparisTarihi) = ?
                                            ) as CustPaid
                                                from            
                                                MusterilerTB m
                                        
                                        """,(yil,yil))
        liste= list()
        for item in result:
            if(item.FOB == None and item.CustPaid == None):
                continue
            else:
                model = MusteriBazindaUretimModel()
                model.musteriAdi = item.MusteriAdi
                if(item.MusteriId == 37):
                    model.marketing = 'Imperial Homes'
                else:
                    model.marketing = item.Marketing
                model.ulkeAdi = item.Ulke
                model.toplam  = item.FOB
                model.toplamCfr = self.__getNoneType(item.FOB) + self.__getNoneType(item.CustPaid)
                liste.append(model)
                
        schema = MusteriBazindaUretimSchema(many=True)
        return schema.dump(liste) 
    
    def getMarketing(self,yil):
        try:
            sipTotal = self.data.getStoreList("""
                                                select 	
                                                sum(su.SatisToplam) as Toplam,
                                                m.Marketing as Marketing
                                            from MusterilerTB m	
                                                inner join SiparislerTB s on s.MusteriID = m.ID
                                                inner join SiparisUrunTB su on su.SiparisNo = s.SiparisNo
                                            where 	
                                                    s.SiparisDurumID = 2 and YEAR(s.SiparisTarihi) = ?
                                            group by
                                                m.Marketing
                                            """,(yil))
            self.marketingNavlun = self.data.getStoreList("""
                                                    select 	
                                                    sum(s.NavlunSatis) + sum(s.DetayTutar_1) + sum(s.DetayTutar_2) + sum(s.DetayTutar_3) + sum(s.DetayTutar_4) as Navlun,
                                                    m.Marketing as Marketing
                                                from MusterilerTB m	
                                                    inner join SiparislerTB s on s.MusteriID = m.ID
                                                where 	
                                                    s.SiparisDurumID=2 and YEAR(s.SiparisTarihi)=?
                                                group by
                                                    m.Marketing
                                                 
                                                 """,(yil))
            
            liste = list()
            for item in sipTotal:
                model = MusteriBazindaUretimModel()
                model.marketing = item.Marketing
                model.toplam = self.__getNoneType(item.Toplam)
                model.toplamCfr = model.toplam + self.__getNoneType(self.__getNavlunUretim(item.Marketing))
                liste.append(model)
            schema = MusteriBazindaUretimSchema(many=True)
            return schema.dump(liste)
        except Exception as e:
            print('getMarketing hata',str(e))
            return False
    
    def getMarketingYukleme(self,year):
        
        result = self.data.getStoreList("""
                                            select 	
                                                sum(su.SatisToplam) as Toplam,
                                                m.Marketing as Marketing
                                            from MusterilerTB m	
                                                inner join SiparislerTB s on s.MusteriID = m.ID
                                                inner join SiparisUrunTB su on su.SiparisNo = s.SiparisNo
                                            where 	
                                                YEAR(s.YuklemeTarihi) = ? and s.SiparisDurumID =3 
                                            group by
                                                m.Marketing

                                        """,(year))
        
        self.navlunYukleme = self.data.getStoreList("""
                                                select 	
                                                    sum(s.NavlunSatis) as Navlun,
                                                    sum(s.DetayTutar_1) as DetayTutar1,
                                                    sum(s.DetayTutar_2) as DetayTutar2,
                                                    sum(s.DetayTutar_3) as DetayTutar3,
                                                    sum(s.DetayTutar_4) as DetayTutar4,
                                                    
                                                    m.Marketing as Marketing
                                                from MusterilerTB m	
                                                    inner join SiparislerTB s on s.MusteriID = m.ID
                                                where 	
                                                    YEAR(s.YuklemeTarihi) = ? and s.SiparisDurumID =3
                                                group by
                                                    m.Marketing                                                
                                             """,(year))
        
        liste = list()
        for item in result:
            model = MarketingModel()
            model.marketing = item.Marketing
            model.fobToplam = item.Toplam
            model.cfrToplam = item.Toplam + self.__getNavlunYukleme(item.Marketing)
            liste.append(model)

            
        schema = MarketingSchema(many=True)
        return schema.dump(liste)
    
    def getMarketingDetail(self,year):
        try:
            result = self.data.getStoreList("""
                                                select            
                                                    m.ID as MusteriId,            
                                                    m.FirmaAdi as MusteriAdi,               
                                                    m.Marketing,
                                                    (select yu.UlkeAdi from YeniTeklif_UlkeTB yu where yu.Id = m.UlkeId ) as Ulke,
                                            
                                    
                                                            
                                        (          
                                        Select Sum(u.SatisToplam) from SiparislerTB s, SiparisUrunTB u where s.SiparisNo=u.SiparisNo and s.SiparisDurumID=3 and s.MusteriID=m.ID and Year(s.YuklemeTarihi)=?
                                            
                                        ) as Toplam 


                                                    from            
                                                    MusterilerTB m
                                            
                                            """,(year))
            self.marketingYuklemeNavlun = self.data.getStoreList("""
                                                select            
                                                    m.ID as MusteriId,            
                                                    m.FirmaAdi as MusteriAdi,               
                                                    m.Marketing,
                                                    (select yu.UlkeAdi from YeniTeklif_UlkeTB yu where yu.Id = m.UlkeId ) as Ulke,
                                            
                                    
                                                            
                                        (          
                                        Select Sum(s.NavlunSatis) from SiparislerTB s where s.SiparisDurumID=3 and s.MusteriID=m.ID and Year(s.YuklemeTarihi)=?  
                                            
                                        ) +
                                        (          
                                        Select Sum(s.DetayTutar_1) from SiparislerTB s where s.SiparisDurumID=3 and s.MusteriID=m.ID and Year(s.YuklemeTarihi)=?
                                            
                                        ) +
                                        (          
                                        Select Sum(s.DetayTutar_2) from SiparislerTB s where s.SiparisDurumID=3 and s.MusteriID=m.ID and Year(s.YuklemeTarihi)=?  
                                            
                                        )+
                                        (          
                                        Select Sum(s.DetayTutar_3) from SiparislerTB s where s.SiparisDurumID=3 and s.MusteriID=m.ID and Year(s.YuklemeTarihi)=? 
                                            
                                        ) +
                                        (          
                                        Select Sum(s.DetayTutar_4) from SiparislerTB s where s.SiparisDurumID=3 and s.MusteriID=m.ID and Year(s.YuklemeTarihi)=?  
                                            
                                        ) as Masraflar


                                                    from            
                                                    MusterilerTB m
                                            
                                            """,(year,year,year,year,year))    
        
            liste = list()
            for item in result:
                model = MarketingAyrintiModel()
                if(item.Toplam == None):
                    continue
                else:
                    model.musteri = item.MusteriAdi
                    model.marketing = item.Marketing
                    model.toplamFob = item.Toplam
                    model.toplamCfr = float(item.Toplam) + float(self.__getMarketingDetailNavlun(item.MusteriId))
                    liste.append(model)
            schema = MarketingAyrintiSchema(many=True)
            return schema.dump(liste)
                
        except Exception as e:
            print('getMarketingDetail hata',str(e))
    
    def mkRaporlarSevkSip(self,year):
        try:
            data = self.data.getStoreList("""
                                    select 
                                        m.FirmaAdi,

                                        (
                                            select sum(su.SatisToplam) from SiparislerTB s, SiparisUrunTB su where  s.MusteriID = m.ID and s.SiparisNo = su.SiparisNo and YEAR(s.SiparisTarihi) = ?
                                        ) as BuYilSiparisler,
                                        (
                                            select sum(su.SatisToplam) from SiparislerTB s, SiparisUrunTB su where s.MusteriID = m.ID and s.SiparisNo = su.SiparisNo and YEAR(s.YuklemeTarihi) = ?
                                        ) as BuYilYuklenenler

                                    from MusterilerTB m
                                    where m.Marketing = 'Mekmar'
                                  
                                  """,(year,year))
            self.yuklenenMusteriMasraf = self.data.getStoreList("""
                                                    select
                                                        m.FirmaAdi,
                                                        sum(s.NavlunSatis) + sum(s.DetayTutar_1) + sum(s.DetayTutar_2) + sum(s.DetayTutar_3) as NavlunvDiger

                                                    from SiparislerTB s
                                                        inner join MusterilerTB m on m.ID = s.MusteriID
                                                    where
                                                        YEAR(s.YuklemeTarihi) = ? and m.Marketing = 'Mekmar'
                                                    group by s.MusteriID,m.FirmaAdi
                                                  """,(year))
            fullyListe = list()
            for item2 in data:
                if(item2.BuYilSiparisler == None and item2.BuYilYuklenenler == None):
                    continue
                else:
                    fullyListe.append(item2)

            liste = list()
            for item in fullyListe:
                model = MkRaporlarSevkSipModel()
                model.siparisfob = self.__getNoneType(item.BuYilSiparisler)
                model.yuklenenfob = self.__getNoneType(item.BuYilYuklenenler)
                model.yuklenenddp = model.yuklenenfob + self.__getNoneType(self.__getYuklenenMasraf(item.FirmaAdi))
                model.musteriadi = item.FirmaAdi
                model.total = model.siparisfob + model.yuklenenddp
                liste.append(model)
            schema = MkRaporlarSevkSipSchema(many=True)
            return schema.dump(liste)
        except Exception as e:
            print('mkRaporlarSevkSip hata',str(e))
            return False
    
    def __getMarketingDetailNavlun(self,musteriId):
        for item in self.marketingYuklemeNavlun:
            if item.MusteriId != musteriId:
                continue
            else:
                if item.Masraflar != None:
                    return item.Masraflar
                else:
                    return 0
    
    def __getNavlunYukleme(self,marketing):
        for item in self.navlunYukleme:
            if item.Marketing != marketing:
                continue
            else:
                return item.Navlun + item.DetayTutar1 + item.DetayTutar2 + item.DetayTutar3 + item.DetayTutar4
    
    def __getNavlunUretim(self,marketing):
        for item in self.marketingNavlun:
            if item.Marketing != marketing:
                continue
            else:
                return item.Navlun       
    
    def __getYuklenenMasraf(self,firma):
        for item in self.yuklenenMusteriMasraf:
            if(item.FirmaAdi == firma):
                return item.NavlunvDiger
    
    def __getNoneType(self,value):
        if(value == None):
            return 0
        else:
            return float(value)
        
    def getMkRaporlariExcelList(self,data):
        try:
            source_path = 'resource_api/raporlar/sablonlar/mkRaporlari.xlsx'
            target_path = 'resource_api/raporlar/dosyalar/mkRaporlari.xlsx'
            shutil.copy2(source_path, target_path)
            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sayfa1')
            
            sayfa.cell(1,column=1,value=datetime.datetime.now().strftime('%Y') + ' YILI BAŞINDAN İTİBAREN ALINAN SİPARİŞLER')
            # sayfa.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
            satir = 3
            
            byCustomerTotal = {
                'fob':0,
                'ddp':0,
            }
            for item in data['byPo']:
                sayfa.cell(satir,column=1,value=item['tarih'])
                sayfa.cell(satir,column=2,value=item['firma'])
                sayfa.cell(satir,column=3,value=item['po'])
                sayfa.cell(satir,column=4,value=item['teslim'])
                sayfa.cell(satir,column=5,value=item['fob'])
                sayfa.cell(satir,column=6,value=item['ddp'])
                byCustomerTotal['fob'] += item['fob']
                byCustomerTotal['ddp'] += item['ddp']
                satir += 1
            sayfa.cell(satir,column=1,value='Toplam')
            sayfa.cell(satir,column=5,value=byCustomerTotal['fob'])
            sayfa.cell(satir,column=6,value=byCustomerTotal['ddp'])

            sayfa2 = kitap.get_sheet_by_name('Sayfa2')
            sayfa2.cell(1,column=1,value=datetime.datetime.now().strftime('%Y') + ' TARİHİ İTİBARİYLE SİPARİŞLER')
            # sayfa2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
            satir2 = 3
            byMarketingTotal = {
                'fob':0,
                'ddp':0,
            }
            for item in data['byMarketing']:
                sayfa2.cell(satir2,column=1,value=item['marketing'])
                sayfa2.cell(satir2,column=2,value=item['toplam'])
                sayfa2.cell(satir2,column=3,value=item['toplamCfr'])
                byMarketingTotal['fob'] += item['toplam']
                byMarketingTotal['ddp'] += item['toplamCfr']
                satir2 += 1
            sayfa2.cell(satir2,column=1,value='Toplam')
            sayfa2.cell(satir2,column=2,value=byMarketingTotal['fob'])
            sayfa2.cell(satir2,column=3,value=byMarketingTotal['ddp'])
            
            sayfa2.cell(1,column=5,value=datetime.datetime.now().strftime('%x') + ' TARİHİ İTİBARİYLE SİPARİŞLER DETAY')
            # sayfa2.merge_cells(start_row=satir2,start_column=1,end_row=satir2,end_column=5)
            satir3=3
            byCustomerTotal = {
                'fob':0,
                'ddp':0,
            }
            for item in data['byCustomer']:
                sayfa2.cell(satir3,column=5,value=item['musteriAdi'])
                sayfa2.cell(satir3,column=6,value=item['marketing'])
                sayfa2.cell(satir3,column=7,value=item['ulkeAdi'])
                sayfa2.cell(satir3,column=8,value=item['toplam'])
                sayfa2.cell(satir3,column=9,value=item['toplamCfr'])
                byCustomerTotal['fob'] += item['toplam']
                byCustomerTotal['ddp'] += item['toplamCfr']
                satir3 += 1
            sayfa2.cell(satir3,column=5,value='Toplam')
            sayfa2.cell(satir3,column=8,value=byCustomerTotal['fob'])
            sayfa2.cell(satir3,column=9,value=byCustomerTotal['ddp'])
            mekmarList = []
            mekmerList = []
            icPiyasaList = []
            imperialHomesList = []

            def siparislerMarketing():
                for item in data['byCustomer']:
                    if(item['marketing'] == 'Mekmar'):
                        mekmarList.append(item)
                    elif item['marketing'] == 'Mekmer':
                        mekmerList.append(item)
                    elif item['marketing'] == 'İç Piyasa':
                        icPiyasaList.append(item)
                    elif item['marketing'] == 'Imperial Homes':
                        imperialHomesList.append(item)
        
            siparislerMarketing()
            
            sayfa2.cell(1,column=11,value='Mekmar')
            satir4 = 3
            mekmarListTotal = {
                'fob':0,
                'ddp':0,
            }
            for item in mekmarList:
                sayfa2.cell(satir4,column=11,value=item['musteriAdi'])
                sayfa2.cell(satir4,column=12,value=item['ulkeAdi'])
                sayfa2.cell(satir4,column=13,value=item['toplam'])
                sayfa2.cell(satir4,column=14,value=item['toplamCfr'])
                mekmarListTotal['fob'] += item['toplam']
                mekmarListTotal['ddp'] += item['toplamCfr']
                satir4 += 1
            sayfa2.cell(satir4,column=11,value='Toplam')
            sayfa2.cell(satir4,column=13,value=mekmarListTotal['fob'])
            sayfa2.cell(satir4,column=14,value=mekmarListTotal['ddp'])
            
            sayfa2.cell(1,column=16,value='İç Piyasa')
            satir5 = 3
            icPiyasaListTotal = {
                'fob':0,
                'ddp':0,    
            }
            for item in icPiyasaList:
                sayfa2.cell(satir5,column=16,value=item['musteriAdi'])
                sayfa2.cell(satir5,column=17,value=item['ulkeAdi'])
                sayfa2.cell(satir5,column=18,value=item['toplam'])
                sayfa2.cell(satir5,column=19,value=item['toplamCfr'])
                icPiyasaListTotal['fob'] += item['toplam']
                icPiyasaListTotal['ddp'] += item['toplamCfr']
                satir5 += 1
            sayfa2.cell(satir5,column=16,value='Toplam')
            sayfa2.cell(satir5,column=18,value=icPiyasaListTotal['fob'])
            sayfa2.cell(satir5,column=19,value=icPiyasaListTotal['ddp'])
            
            sayfa2.cell(1,column=21,value='Mekmer')
            satir6 = 3
            mekmerTotal = {
                'fob':0,
                'ddp':0,    
            }
            for item in mekmerList:
                sayfa2.cell(satir6,column=21,value=item['musteriAdi'])
                sayfa2.cell(satir6,column=22,value=item['ulkeAdi'])
                sayfa2.cell(satir6,column=23,value=item['toplam'])
                sayfa2.cell(satir6,column=24,value=item['toplamCfr'])
                mekmerTotal['fob'] += item['toplam']
                mekmerTotal['ddp'] += item['toplamCfr']
                satir6 += 1
            sayfa2.cell(satir6,column=21,value='Toplam')
            sayfa2.cell(satir6,column=23,value=mekmerTotal['fob'])
            sayfa2.cell(satir6,column=24,value=mekmerTotal['ddp'])
            
            sayfa2.cell(1,column=26,value='İmperial Homes')
            satir7 = 3
            imperialHomesTotal = {
                'fob':0,
                'ddp':0,    
            }
            for item in imperialHomesList:
                sayfa2.cell(satir7,column=26,value=item['musteriAdi'])
                sayfa2.cell(satir7,column=27,value=item['ulkeAdi'])
                sayfa2.cell(satir7,column=28,value=item['toplam'])
                sayfa2.cell(satir7,column=29,value=item['toplamCfr'])
                imperialHomesTotal['fob'] += item['toplam']
                imperialHomesTotal['ddp'] += item['toplamCfr']
                satir7 += 1
            sayfa2.cell(satir7,column=26,value='Toplam')
            sayfa2.cell(satir7,column=28,value=imperialHomesTotal['fob'])
            sayfa2.cell(satir7,column=29,value=imperialHomesTotal['ddp'])

            sayfa3 = kitap.get_sheet_by_name('Sayfa3')
            sayfa3.cell(1,column=1,value=  '1-1-2023' + ' ' + datetime.datetime.now().strftime('%x') + ' ARASI YÜKLEMELER')
            satir8 = 3
            byMarketingYuklemeTotal = {
                'fob':0,
                'ddp':0,
            }
            for item in data['byMarketingYukleme']:
                sayfa3.cell(satir8,column=1,value=item['marketing'])
                sayfa3.cell(satir8,column=2,value=item['fobToplam'])
                sayfa3.cell(satir8,column=3,value=item['cfrToplam'])
                byMarketingYuklemeTotal['fob'] += item['fobToplam']
                byMarketingYuklemeTotal['ddp'] += item['cfrToplam']
                satir8 += 1
            sayfa3.cell(satir8,column=1,value='Toplam')
            sayfa3.cell(satir8,column=2,value=byMarketingYuklemeTotal['fob'])
            sayfa3.cell(satir8,column=3,value=byMarketingYuklemeTotal['ddp'])
            
            sayfa4 = kitap.get_sheet_by_name('Sayfa4')
            satir13 = 2
            for item in data['byCustomerOrder']:
                sayfa4.cell(satir13,column=1,value=item['musteri'])
                sayfa4.cell(satir13,column=2,value=item['ulkeAdi'])
                sayfa4.cell(satir13,column=3,value=item['temsilci'])
                
                if item['Toplam'] == None:
                    sayfa4.cell(satir13,column=4,value=0)
                else:
                    sayfa4.cell(satir13,column=4,value=item['Toplam'])
                
                
                
                if item['BuYilUretim'] == None:
                    sayfa4.cell(satir13,column=5,value=0) 
                else:
                    sayfa4.cell(satir13,column=5,value=item['BuYilUretim']) 

                if item['BuYilSevkiyat'] == None:
                    sayfa4.cell(satir13,column=6,value=0) 
                else:
                    sayfa4.cell(satir13,column=6,value=item['BuYilSevkiyat']) 
                
                

                if item['GecenYil'] == None:
                    sayfa4.cell(satir13,column=7,value=0) 
                else:
                    sayfa4.cell(satir13,column=7,value=item['GecenYil']) 

                if item['OncekiYil'] == None:
                    sayfa4.cell(satir13,column=8,value=0) 
                else:
                    sayfa4.cell(satir13,column=8,value=item['OncekiYil'])
                    
                if item['OnDokuzYili'] == None:
                    sayfa4.cell(satir13,column=9,value=0) 
                else:
                    sayfa4.cell(satir13,column=9,value=item['OnDokuzYili']) 
                    
                if item['OnSekizYili'] == None:
                    sayfa4.cell(satir13,column=10,value=0) 
                else:
                    sayfa4.cell(satir13,column=10,value=item['OnSekizYili'])                 
                
                if item['OnYediYili'] == None:
                    sayfa4.cell(satir13,column=11,value=0) 
                else:
                    sayfa4.cell(satir13,column=11,value=item['OnYediYili']) 
                
                if item['OnAltiYili'] == None:
                    sayfa4.cell(satir13,column=12,value=0) 
                else:
                    sayfa4.cell(satir13,column=12,value=item['OnAltiYili']) 
                    
                if item['OnBesYili'] == None:
                    sayfa4.cell(satir13,column=13,value=0) 
                else:
                    sayfa4.cell(satir13,column=13,value=item['OnBesYili']) 
                
                
                if item['OnDortYili'] == None:
                    sayfa4.cell(satir13,column=14,value=0) 
                else:
                    sayfa4.cell(satir13,column=14,value=item['OnDortYili']) 
                
                if item['OnUcYili'] == None:
                    sayfa4.cell(satir13,column=15,value=0) 
                else:
                    sayfa4.cell(satir13,column=15,value=item['OnUcYili']) 
                
                if item['OnUcYiliOncesi'] == None:
                    sayfa4.cell(satir13,column=16,value=0) 
                else:
                    sayfa4.cell(satir13,column=16,value=item['OnUcYiliOncesi'])
                satir13 += 1
            
            
            mekmarYuklemeList = []
            mekmerYuklemeList = []
            icPiyasaYuklemeList = []
            imperialHomesYuklemeList = []

            def yuklemelerMarketing():
                for item in data['byMarketingDetayYukleme']:
                    if(item['marketing'] == 'Mekmar'):
                        mekmarYuklemeList.append(item)
                    elif item['marketing'] == 'Mekmer':
                        mekmerYuklemeList.append(item)
                    elif item['marketing'] == 'İç Piyasa':
                        icPiyasaYuklemeList.append(item)
                    elif item['marketing'] == 'Imperial Homes':
                        imperialHomesYuklemeList.append(item)
            yuklemelerMarketing()
            
            sayfa3.cell(1,column=5,value=  'Mekmar')
            satir9 = 3
            mekmarYuklemeListTotal = {
                'fob':0,
                'ddp':0,
            }
            for item in mekmarYuklemeList:
                sayfa3.cell(satir9,column=5,value=item['musteri'])
                sayfa3.cell(satir9,column=6,value=item['toplamFob'])
                sayfa3.cell(satir9,column=7,value=item['toplamCfr'])
                mekmarYuklemeListTotal['fob'] += item['toplamFob']
                mekmarYuklemeListTotal['ddp'] += item['toplamCfr']
                satir9 += 1
            sayfa3.cell(satir9,column=5,value='Toplam')
            sayfa3.cell(satir9,column=6,value=mekmarYuklemeListTotal['fob'])
            sayfa3.cell(satir9,column=7,value=mekmarYuklemeListTotal['ddp'])
            
            sayfa3.cell(1,column=9,value=  'İç Piyasa')
            satir10 = 3
            icpiyasaYuklemeListTotal = {
                'fob':0,
                'ddp':0,
            }
            for item in icPiyasaYuklemeList:
                sayfa3.cell(satir10,column=9,value=item['musteri'])
                sayfa3.cell(satir10,column=10,value=item['toplamFob'])
                sayfa3.cell(satir10,column=11,value=item['toplamCfr'])
                icpiyasaYuklemeListTotal['fob'] += item['toplamFob']
                icpiyasaYuklemeListTotal['ddp'] += item['toplamCfr']
                satir10 += 1
            sayfa3.cell(satir10,column=9,value='Toplam')
            sayfa3.cell(satir10,column=10,value=icpiyasaYuklemeListTotal['fob'])
            sayfa3.cell(satir10,column=11,value=icpiyasaYuklemeListTotal['ddp'])
            
            sayfa3.cell(1,column=13,value=  'Mekmer')
            satir11 = 3
            mekmerYuklemeListTotal = {
                'fob':0,
                'ddp':0,
            }
            for item in mekmerYuklemeList:
                sayfa3.cell(satir11,column=13,value=item['musteri'])
                sayfa3.cell(satir11,column=14,value=item['toplamFob'])
                sayfa3.cell(satir11,column=15,value=item['toplamCfr'])
                mekmerYuklemeListTotal['fob'] += item['toplamFob']
                mekmerYuklemeListTotal['ddp'] += item['toplamCfr']
                satir11 += 1
            sayfa3.cell(satir11,column=13,value='Toplam')
            sayfa3.cell(satir11,column=14,value=mekmerYuklemeListTotal['fob'])
            sayfa3.cell(satir11,column=15,value=mekmerYuklemeListTotal['ddp'])
            
            sayfa3.cell(1,column=17,value=  'İmperial Homes')
            satir12 = 3
            imperialHomesYuklemeListTotal = {
                'fob':0,
                'ddp':0,
            }
            for item in imperialHomesYuklemeList:
                sayfa3.cell(satir12,column=17,value=item['musteri'])
                sayfa3.cell(satir12,column=18,value=item['toplamFob'])
                sayfa3.cell(satir12,column=19,value=item['toplamCfr'])
                imperialHomesYuklemeListTotal['fob'] += item['toplamFob']
                imperialHomesYuklemeListTotal['ddp'] += item['toplamCfr']
                satir12 += 1
            sayfa3.cell(satir12,column=17,value='Toplam')
            sayfa3.cell(satir12,column=18,value=imperialHomesYuklemeListTotal['fob'])
            sayfa3.cell(satir12,column=19,value=imperialHomesYuklemeListTotal['ddp'])
            
            sayfa5 = kitap.get_sheet_by_name('Sayfa5')
            satir15=2
            for item in data['byYuklemevSiparisler']:
                sayfa5.cell(satir15,column=1,value=item['musteriadi'])
                sayfa5.cell(satir15,column=2,value=item['siparisfob'])
                sayfa5.cell(satir15,column=3,value=item['yuklenenddp'])
                sayfa5.cell(satir15,column=4,value=item['total'])
                satir15 += 1
            

                
            
            
            kitap.save(target_path)
            kitap.close()
            return True

            
            
        except Exception as e:
            print('getMkRaporlariExcelList hata',str(e))
            return False